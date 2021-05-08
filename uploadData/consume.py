from openpyxl import load_workbook
import pymysql

conn = pymysql.connect(host='127.0.0.1', user='root', password='guyanjie0908', database='graduate')
cursor = conn.cursor()
file = load_workbook('f:/消费数据信息.xlsx', 'r')
store = set()
storeNames = []
i = 0
for sheet in file.sheetnames:
    for row in file[sheet].rows:
        i += 1
        if i == 1:
            continue
        store_name = row[3].value.strip()
        sid = str(row[0].value)
        money = float(row[4].value)
        dateTime = str(row[5].value)
        mode = str(row[7].value)
        if store_name not in store:
            sql = "insert into store(store_name) values ('{}')".format(store_name)
            cursor.execute(sql)
            conn.commit()
            sql = 'select * from store'
            cursor.execute(sql)
            storeNames = cursor.fetchall()
        for storeN in storeNames:
            if storeN[1] == store_name:
                sql = "insert into consume(sid,execution_time,money,store_id,mode) values ('{}','{}',{},'{}','{}')"\
                    .format(sid, dateTime, money, storeN[0], mode)
                cursor.execute(sql)
                conn.commit()
                break
        store.add(store_name)
        i += 1
        print(i)
